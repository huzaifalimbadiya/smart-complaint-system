from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO

from complaints.models import Complaint, Category, AdminResponse
from complaints.forms import ComplaintUpdateForm, AdminResponseForm, ComplaintSearchForm
from complaints.decorators import admin_required


@admin_required
def admin_dashboard(request):
    """Main admin dashboard with statistics and charts"""
    
    # Overall statistics
    total_complaints = Complaint.objects.count()
    pending_complaints = Complaint.objects.filter(status='pending').count()
    in_progress_complaints = Complaint.objects.filter(status='in_progress').count()
    resolved_complaints = Complaint.objects.filter(status='resolved').count()
    rejected_complaints = Complaint.objects.filter(status='rejected').count()
    
    # Category-wise statistics
    category_stats = Complaint.objects.values(
        'final_category__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Recent complaints
    recent_complaints = Complaint.objects.select_related(
        'user', 'final_category'
    ).order_by('-created_at')[:10]
    
    # Monthly trend (last 6 months)
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_data = []
    for i in range(6):
        month_start = timezone.now() - timedelta(days=30 * (5 - i))
        month_end = timezone.now() - timedelta(days=30 * (4 - i))
        count = Complaint.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end
        ).count()
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'count': count
        })
    
    # Convert chart data to JSON for JavaScript
    category_labels = [stat['final_category__name'] or 'Uncategorized' for stat in category_stats]
    category_counts = [stat['count'] for stat in category_stats]
    monthly_labels = [data['month'] for data in monthly_data]
    monthly_counts = [data['count'] for data in monthly_data]
    
    context = {
        'total_complaints': total_complaints,
        'pending_complaints': pending_complaints,
        'in_progress_complaints': in_progress_complaints,
        'resolved_complaints': resolved_complaints,
        'rejected_complaints': rejected_complaints,
        'category_stats': category_stats,
        'recent_complaints': recent_complaints,
        'monthly_data': monthly_data,
        # JSON data for charts
        'category_labels_json': json.dumps(category_labels),
        'category_counts_json': json.dumps(category_counts),
        'monthly_labels_json': json.dumps(monthly_labels),
        'monthly_counts_json': json.dumps(monthly_counts),
    }
    
    return render(request, 'dashboard/admin_dashboard.html', context)


@admin_required
def manage_complaints(request):
    """View for managing all complaints with filtering and sorting"""
    
    # Get all complaints
    complaints = Complaint.objects.select_related(
        'user', 'predicted_category', 'final_category'
    ).all()
    
    # Apply search and filters
    form = ComplaintSearchForm(request.GET)
    if form.is_valid():
        search = form.cleaned_data.get('search')
        status = form.cleaned_data.get('status')
        category = form.cleaned_data.get('category')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        
        if search:
            complaints = complaints.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(area__icontains=search) |
                Q(user__username__icontains=search)
            )
        
        if status:
            complaints = complaints.filter(status=status)
        
        if category:
            complaints = complaints.filter(final_category=category)
        
        if date_from:
            complaints = complaints.filter(created_at__gte=date_from)
        
        if date_to:
            # Add one day to include the entire end date
            complaints = complaints.filter(created_at__lte=date_to)
    
    # Sorting
    sort_by = request.GET.get('sort', '-created_at')
    complaints = complaints.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(complaints, 20)  # 20 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'complaints': page_obj,
        'form': form,
        'total_complaints': complaints.count(),
    }
    
    return render(request, 'dashboard/manage_complaints.html', context)


@admin_required
def update_complaint(request, complaint_id):
    """View for updating complaint status and details"""
    
    complaint = get_object_or_404(
        Complaint.objects.select_related('user', 'predicted_category', 'final_category')
        .prefetch_related('responses__admin'),
        id=complaint_id
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            update_form = ComplaintUpdateForm(request.POST, instance=complaint)
            if update_form.is_valid():
                update_form.save()
                messages.success(request, 'Complaint updated successfully!')
                return redirect('dashboard:update_complaint', complaint_id=complaint_id)
        
        elif action == 'add_response':
            response_form = AdminResponseForm(request.POST)
            if response_form.is_valid():
                response = response_form.save(commit=False)
                response.complaint = complaint
                response.admin = request.user
                response.save()
                messages.success(request, 'Response added successfully!')
                return redirect('dashboard:update_complaint', complaint_id=complaint_id)
        
        elif action == 'delete':
            complaint.delete()
            messages.success(request, 'Complaint deleted successfully!')
            return redirect('dashboard:manage_complaints')
    
    else:
        update_form = ComplaintUpdateForm(instance=complaint)
        response_form = AdminResponseForm()
    
    context = {
        'complaint': complaint,
        'update_form': update_form,
        'response_form': response_form,
        'responses': complaint.responses.all().order_by('created_at')
    }
    
    return render(request, 'dashboard/update_complaint.html', context)


@admin_required
def export_excel(request):
    """Export complaints to Excel"""
    
    # Get filtered complaints
    complaints = Complaint.objects.select_related(
        'user', 'final_category'
    ).all()
    
    # Apply filters if any
    status = request.GET.get('status')
    category = request.GET.get('category')
    
    if status:
        complaints = complaints.filter(status=status)
    if category:
        complaints = complaints.filter(final_category_id=category)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints Report"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['ID', 'Title', 'User', 'Category', 'Area', 'Status', 'Created Date', 'Updated Date']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for complaint in complaints:
        ws.append([
            complaint.id,
            complaint.title,
            complaint.user.username,
            complaint.get_category().name if complaint.get_category() else 'Uncategorized',
            complaint.area,
            complaint.get_status_display(),
            complaint.created_at.strftime('%Y-%m-%d %H:%M'),
            complaint.updated_at.strftime('%Y-%m-%d %H:%M'),
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=complaints_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@admin_required
def export_pdf(request):
    """Export complaints to PDF"""
    
    # Get filtered complaints
    complaints = Complaint.objects.select_related(
        'user', 'final_category'
    ).all()[:100]  # Limit to 100 for PDF
    
    # Apply filters if any
    status = request.GET.get('status')
    category = request.GET.get('category')
    
    if status:
        complaints = complaints.filter(status=status)
    if category:
        complaints = complaints.filter(final_category_id=category)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph("Complaints Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Report info
    info_text = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}<br/>Total Complaints: {complaints.count()}"
    info = Paragraph(info_text, styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Table data
    data = [['ID', 'Title', 'User', 'Category', 'Status', 'Date']]
    
    for complaint in complaints:
        data.append([
            str(complaint.id),
            complaint.title[:30] + '...' if len(complaint.title) > 30 else complaint.title,
            complaint.user.username,
            complaint.get_category().name if complaint.get_category() else 'N/A',
            complaint.get_status_display(),
            complaint.created_at.strftime('%Y-%m-%d'),
        ])
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
    
    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Create response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=complaints_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response
